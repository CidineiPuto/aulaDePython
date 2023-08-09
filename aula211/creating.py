# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # Planilha
# Criando planilha

sheet_name = "Minha planilha"  # Nome da planilha
# Tem duas maneiras de colocar a sua planilha para você mexer nela, e não em
# outra, a primeira é colocar "0" na frente do nome da planilha.
# A nossa planilha será a primeira assim "workbook.create_sheet(sheet_name, 0)"
workbook.create_sheet(sheet_name)
# ou colocar o nome da planilha na frente da variável, exemplo;
worksheet: Worksheet = workbook[sheet_name]
# print(workbook.sheetnames)  # nome das planilhas.

# Remover uma planilha

workbook.remove(workbook["Sheet"])

# criando os cabeçalhos

worksheet.cell(1, 1, "Nome")  # Linha 1, coluna 1, terá o cabeçalho nome
worksheet.cell(1, 2, "Idade")  # Linha 1, coluna 2, terá o cabeçalho Idade
worksheet.cell(1, 3, "Nota")  # Linha 1, coluna 3, terá o cabeçalho nota

students = [
    # nome  idade nota
    ["João", 14, 5.5],
    ["Maria", 13, 9.7],
    ["Luiz", 15, 8.8],
    ["Alberto", 16, 10],
]


"""
Forma mais díficil de fazer.
"""

# Começa de 2 para não alterar o cabeçalho
# for i in range(2, 10):  # linhas
#     for j in range(1, 4 + 1):  # Coluna
#         print("Linha", i, "Coluna", j)

# for i, student_row in enumerate(students, start=2):  # Começa do 2
#     for j, student_column in enumerate(student_row, start=1):  # Começa do 1
#         worksheet.cell(i, j, student_column)

# workbook.save(WORKBOOK_PATH)

"""
Forma mais fácil de fazer.
"""
for student in students:
    worksheet.append(student)  # Irá adicionar um iterável na linha depois, e
    # como só tem uma linha, irá ir adicionando de forma igual na maneira
    # díficil

workbook.save(WORKBOOK_PATH)
